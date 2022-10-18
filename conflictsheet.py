import bisect
import enum
import io
import itertools
import operator
import pickle
import re
import string
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor
from copy import copy
from random import choice, choices, random, sample, shuffle
from operator import attrgetter
from functools import partial, reduce
from typing import *

import numpy as np
import openpyxl
import pandas as pd
import portion as P
import polars as pl
import texttable
from cachetools import LFUCache, cachedmethod
from cachetools.keys import hashkey
from cachetools.func import lfu_cache

Portion = NewType('Portion', P.Interval)

from pg import CURRENT_SEASON, MAX_PG_NAME_LEN, PG, PGPlaying

MISSING_PG = '??????????'
ANY_SLOT = frozenset({1, 2, 3, 4, 5, 6})
FIRST_HALF = frozenset({1, 2, 3})
SECOND_HALF = frozenset({4, 5, 6})
ANY_FLAG = frozenset()
ANY_FREQ = frozenset()

from util import PLAYING_FLAGS, SORT_PROD, build_flag_expr as has_any_flags


_pf_bit_mapping = {pf: len(PLAYING_FLAGS) - 1 - PLAYING_FLAGS.index(pf) for pf in ('^', '?', 'MDG')}


def _pf_bit(pf):
    return _pf_bit_mapping[pf]


Q_FLAG = frozenset((len(PLAYING_FLAGS) - PLAYING_FLAGS.index('?'),))
U_FLAG = frozenset((len(PLAYING_FLAGS) - PLAYING_FLAGS.index('^'),))
QU_FLAGS = Q_FLAG | U_FLAG
ALL_FLAGS_BUT_GUESS = frozenset(range(0, len(PLAYING_FLAGS))) - Q_FLAG
ALL_FLAGS_BUT_UNCERTAIN = ALL_FLAGS_BUT_GUESS - U_FLAG

FLAG_INFO = OrderedDict(
    {
        'C': 'non-car for a non-restored car (NCFAC), or a car game for a boat. Shows up as `car` or `boat` in output.',
        'T': '3+ multiprizer for trips, or in the one instance of the all-Plinko show, Plinko for a trip.',
        '&': "any playing for 2 or more cars that isn't It's Optional or Triple Play. Shows up as `cars` in output.",
        '*': 'A playing with rule changes just for that playing (such as a Big Money Week version of a casher, or 35th Anniversary Plinko), a LMAD Mashup game, or in two syndicated cases, a "vintage car reproduction". Note for MDS shows, an increase on the top prize on cashers was so common that these instances are not denoted with this flag.',
        '@': 'A playing for a restored car, or a 4-prizer played for LA sports season tickets.',
        'R': 'A playing for really unusual prize(s).',
        '$': "Mainly a non-cash game played for cash, hence the use of the dollar sign here. In a couple instances instead, a car game for a trailer, or in the one instance of the all-Plinko show, Plinko for 2 regular prizes.",
        '^': 'The slotting of the game is uncertain. Some old records are incomplete and slotted by best guess.',
        '?': 'The identity of the game is uncertain. Some old records are incomplete and this is a best guess on what the game was. Most often it is known which set of two or three games occurred within a small subset of shows, with no further certainty.',
        'M': 'This was the Million Dollar Game in a Drew MDS (primetime) in 2008. Shows up as `MDG` in output.',
    }
)


@enum.unique
class SlotCertainty(enum.IntFlag):
    CERTAIN = 0
    SLOT = 2 ** _pf_bit('^')
    GAME = 2 ** _pf_bit('?')


UNAIRED_DUPLICATES = ['0013R', '58XXD']


def _dt_convert(dt):
    try:
        return pd.to_datetime(dt, format='%m/%d/%Y')
    except ValueError:
        return np.nan


def _pg_convert(pg_str):
    if not pg_str:
        return PGPlaying(str(PG._UNKNOWN), 1)
    if pg_str == '-':
        return np.nan
    else:
        if unc := pg_str.startswith('*') and pg_str.endswith('*'):
            pg_str = pg_str[1:-1]
        return PGPlaying(pg_str, 2 ** _pf_bit('?') if unc else 0)


conv_dict = {'INT. DATE': _dt_convert, 'AIRDATE': _dt_convert}
conv_dict.update({f'PG{i}': _pg_convert for i in range(1, 7)})


pct_chance = lambda pct: random() < pct / 100


class ConflictSheet:
    _MAX_CACHE = 64
    _CACHE_GETTER = attrgetter('cache')

    def __init__(
        self, load_func: Callable[[], io.BytesIO], save_func: Callable[[io.BytesIO], None], override_pickle: bool = False
    ):
        self.load_func = load_func
        self.save_func = save_func
        self.cache = LFUCache(self._MAX_CACHE)
        try:
            if override_pickle:
                raise ValueError
            self.load_pickle()
            self.excel_fp = None
        except:
            self.load_excel()
            self._df_dict = {}
            self.initialize()

    def get(self, time: str):
        return self._df_dict[time]

    def is_ready(self):
        return bool(self._df_dict)

    def gen_sheet(self, pg_playings: Sequence[PGPlaying], endpoints: Portion, seasonText: str, time: str):
        assert len(pg_playings) == 6
        pgs = [pgp.pg for pgp in pg_playings]

        header = [seasonText, 'SLF', 'TOT', 'SL%']
        header[1:1] = [pg.sheet_abbr if pg else '' for pg in pgs]

        ttable = texttable.Texttable(max_width=0)
        ttable.header(header)
        ttable.set_cols_align('lrrrrrrrrr')
        ttable.set_header_align('c' + 'r' * 9)
        ttable.set_cols_dtype('tiiiiiiiif')
        ttable.set_precision(1)

        sub_slots_df = self.endpoint_sub(endpoints, time, q=self.slot_table(time, 'S'))
        # permutation must be done to match both possible orders of pg1/pg2, so do the more costly operation once
        pg_pair_perms = set(itertools.permutations([str(pg) for pg in pgs], 2))

        pg_pair_dict = Counter()
        for r in self.endpoint_sub(endpoints, time).select(pl.col('^PG[1-6]_p$')).collect().rows():
            for pg1, pg2 in set(itertools.combinations(r, 2)) & pg_pair_perms:
                pg_pair_dict[pg1, pg2] += 1
                pg_pair_dict[pg2, pg1] += 1

        nums = np.zeros((6, 8), dtype=int)
        hundo = False
        for i, pgp in enumerate(pg_playings, 1):
            row = [pgp]
            pg = pgp.pg

            pg_conf = [pg_pair_dict[str(pg), str(pgs[j - 1])] if i != j else '-' for j in range(1, 7)]
            nums[i - 1, :6] = [i if type(i) is int else 0 for pgc in pg_conf]
            row.extend(pg_conf)

            ssd_pg = (
                sub_slots_df.filter((pl.col('PG') == str(pg)) & (pl.col('flag') < SlotCertainty.SLOT))
                .select('^PG\d$')
                .collect()
                .sum()
            )

            nums[i - 1, 6] = slf = ssd_pg[0, i - 1]
            nums[i - 1, 7] = tot = sum(ssd_pg.row(0))
            hundo |= slf == tot
            row.extend([slf, tot, 100.0 * slf / tot if tot else '----'])

            ttable.add_row(row)

        ttable.set_cols_width(
            [max(MAX_PG_NAME_LEN, len(seasonText), max(len(str(pgp)) for pgp in pg_playings))]
            + [max(3, 1 + int(np.log10(n.max()))) for n in nums.T]
            + [5 if hundo else 4]
        )

        return ttable.draw()

    def _pick_slot(self, pg, initial_slots):
        slots = copy(initial_slots)
        if pg in PG.partition_table['NO_OPENING_ACT']:
            slots -= {1, 2}
        elif pg in PG.partition_table['NO_FIRST']:
            slots -= {1}
        if not slots:
            slots = copy(initial_slots)
        return choice(tuple(slots))

    def _pick_game(self, pg_group):
        t = tuple(pg_group)

    def gen_lineup(self, pg_sample: Set[PG], half_hour: bool = False):
        nPGs = [None] * (3 if half_hour else 6)
        non_car = [None] * (3 if half_hour else 6)
        unused_slots = set(range(1, 7))
        unused_fees = ['GP', 'SP']
        halves = [{1, 2, 3}, {4, 5, 6}]
        unused_fee_halves = [0, 1]

        # decide on cash or no cash.
        casher = choice(tuple(pg_sample & PG.partition_table['CASH'])) if pct_chance(70) else None
        cash_type = None if not casher else 'SP' if casher in PG.partition_table['SP/CASH'] else 'GP'
        if casher:
            slot = self._pick_slot(casher, unused_slots)
            nPGs[slot - 1] = casher
            unused_slots.remove(slot)
            pg_sample -= PG.partition_table[f'{cash_type}']
            if casher in PG.partition_table['BAILOUT']:
                pg_sample -= PG.partition_table['BAILOUT']
            unused_fees.remove(cash_type)
            unused_fee_halves.remove((slot - 1) // 3)
            if slot == 1 and pct_chance(98):
                pg_sample.discard(PG.GoldenRoad)

        # now decide cars.
        total_car_count = 3 if pct_chance(2) else 2
        for ttc in range(total_car_count):
            do_non_car = pct_chance(4)
            car_sample = pg_sample & (PG.partition_table['NON-CAR'] if do_non_car else PG.partition_table['CAR'])

            respect_halves = pct_chance(93)
            slot_choices = copy(unused_slots)
            if ttc == 0:
                slot_choices &= halves[0]
                if 0 not in unused_fee_halves and respect_halves:
                    car_sample -= PG.partition_table['FEE']
            if ttc == 1:
                slot_choices &= halves[1]
                if 1 not in unused_fee_halves and respect_halves:
                    car_sample -= PG.partition_table['FEE']
            if nPGs[2] in PG.partition_table['CAR'] and pct_chance(99.5):
                slot_choices -= {4}

            car = choice(tuple(car_sample))

            if car == PG.GoldenRoad and not nPGs[0] and pct_chance(95):
                slot = 1
            else:
                slot = self._pick_slot(car, slot_choices)
            nPGs[slot - 1] = car
            unused_slots.remove(slot)
            non_car[slot - 1] = do_non_car
            car_type = 'SP' if car in PG.partition_table['SP/CAR'] else 'GP' if car in PG.partition_table['GP/CAR'] else None
            if car_type:
                pg_sample -= PG.partition_table[f'{car_type}']
                if respect_halves:
                    unused_fees.remove(car_type)
                    unused_fee_halves.remove((slot - 1) // 3)
            else:
                pg_sample.remove(car)

        if len(unused_fees) == 2 or pct_chance(95):
            shuffle(unused_fees)
            shuffle(unused_fee_halves)
            # fill in unused fees with regular fees, respecting halves most of the time.
            for fee, half in zip(unused_fees, unused_fee_halves):
                sample = pg_sample & PG.partition_table[f'REG. {fee}']
                if sample:
                    mp = choice(tuple(sample))
                    respect_halves = pct_chance(95)
                    slot = self._pick_slot(mp, unused_slots & halves[half] if respect_halves else unused_slots)
                    nPGs[slot - 1] = mp
                    unused_slots.remove(slot)
                    pg_sample -= PG.partition_table[f'REG. {fee}']

        # decide on a 4 prizer or not.
        do_4p = (
            bool(pg_sample & PG.partition_table['4 PRIZER'])
            and not (set(nPGs) & {PG.MoreOrLess, PG.FortuneHunter})
            and pct_chance(27)
        )
        if do_4p:
            mp = choice(tuple(pg_sample & PG.partition_table['4 PRIZER']))
            slot = self._pick_slot(mp, unused_slots)
            nPGs[slot - 1] = mp
            unused_slots.remove(slot)
            pg_sample -= PG.partition_table['4 PRIZER']

        # decide on a 3 prizer or not.
        do_3p = bool(pg_sample & PG.partition_table['3 PRIZER']) and pct_chance(3 if do_4p else 30)
        if do_3p:
            mp = choice(tuple(pg_sample & PG.partition_table['3 PRIZER']))
            slot = self._pick_slot(mp, unused_slots)
            nPGs[slot - 1] = mp
            unused_slots.remove(slot)
            pg_sample -= PG.partition_table['3 PRIZER']

        # decide on a 2 prizer or not.
        do_2p = bool(pg_sample & PG.partition_table['2 PRIZER']) and pct_chance(
            0.5 if do_4p and do_3p else 40 if do_3p else 75
        )
        if do_2p:
            mp = choice(tuple(pg_sample & PG.partition_table['2 PRIZER']))
            slot = self._pick_slot(mp, unused_slots)
            nPGs[slot - 1] = mp
            unused_slots.remove(slot)
            pg_sample -= PG.partition_table['2 PRIZER']

        if unused_slots:
            # decide on a 1+ prizer or not.
            do_1plusp = bool(pg_sample & PG.partition_table['1+ PRIZER']) and pct_chance(
                0.25 if (do_4p + do_3p + do_2p >= 2) else 10 if (do_4p + do_3p >= 1) else 25 if do_2p and do_3p else 65
            )
            if do_1plusp:
                mp = choice(tuple(pg_sample & PG.partition_table['1+ PRIZER']))
                slot = self._pick_slot(mp, unused_slots)
                nPGs[slot - 1] = mp
                unused_slots.remove(slot)
                pg_sample -= PG.partition_table['1+ PRIZER']

        while unused_slots:
            # fill out remainder of lineup with 1 prizers.
            mp = choice(tuple(pg_sample & PG.partition_table['1 PRIZER']))
            slot = self._pick_slot(mp, unused_slots)
            nPGs[slot - 1] = mp
            unused_slots.remove(slot)
            pg_sample.remove(mp)

        return nPGs, non_car

    @cachedmethod(_CACHE_GETTER, key=partial(hashkey, 'ep_sub'))
    def endpoint_sub(self, endpoints: Portion, time: str, *, q: Optional[pl.LazyFrame] = None):
        if q is None:
            q = self._df_dict[time].lazy()
        if not endpoints:
            return q
        elif type(endpoints.lower) is not int:
            return q.filter(pl.col('AIRDATE').is_in(pl.date_range(endpoints.lower, endpoints.upper, '1d')))
        elif time != 'primetime':
            return q.filter(pl.col('S').is_in(list(P.iterate(endpoints, step=1))))
        else:
            return q

    @cachedmethod(_CACHE_GETTER, key=partial(hashkey, 'lineup'))
    def lineup_query(
        self,
        endpoints: Portion,
        time: str,
        logic: str,
        psff_quads: Sequence[tuple[frozenset[PG], frozenset[int | str], Optional[frozenset[int]], Optional[frozenset[int]]]],
    ):
        q = self.endpoint_sub(endpoints, time)
        exprs = []

        for pgs, slots, flags, freqs in psff_quads:
            e = [pl.col(f'PG{s}_p').is_in([str(pg) for pg in pgs]) for s in slots]

            if flags:
                e = [ee & has_any_flags(f'PG{s}_f', frozenset(flags)) for ee, s in zip(e, slots)]

            if freqs:
                exprs.append(sum([ee.cast(pl.UInt8) for ee in e]).is_in(tuple(freqs)))
            else:
                exprs.append(pl.any(e))

        if logic == 'all':
            total_expr = pl.all(exprs)
        elif logic == 'any':
            total_expr = pl.any(exprs)
        else:
            l = locals()
            l |= {letter: fe for fe, letter in zip(exprs, string.ascii_uppercase)}
            total_expr = eval(re.sub('([A-Z])', r'(\1)', logic))

        return q.filter(total_expr)

    @cachedmethod(_CACHE_GETTER, key=partial(hashkey, 'cc'))
    def concurrence_query(
        self, endpoints: Portion, time: str, pgQueries: Tuple[PG], pgFlags: tuple[Optional[frozenset[int]]]
    ):
        q = self.endpoint_sub(endpoints, time)

        pgs = [str(pg) for pg in pgQueries]
        pg_end_label = 3 if time == 'syndicated' else 6

        if not all(pgFlags):
            q = q.with_column(pl.concat_list(pl.col(f'^PG[1-{pg_end_label}]_p$')).alias('PG_a'))

        for pg, pgf in zip(pgs, pgFlags):
            if pgf:
                exprs = [(pl.col(f'PG{i}_p') == pg) & (has_any_flags(f'PG{i}_f', pgf)) for i in range(1, pg_end_label + 1)]
                q = q.filter(pl.any(exprs))
            else:
                q = q.filter(pl.col('PG_a').arr.contains(pg))

        return q

    @cachedmethod(_CACHE_GETTER, key=partial(hashkey, 'slots'))
    def slot_table(self, time: str, by: Optional[str] = None):
        q = self._df_dict[time].lazy()

        vc_subset = [[f'PG{i}_p', f'PG{i}_f'] for i in range(1, 4 if time == 'syndicated' else 7)]
        if by:
            for vc in vc_subset:
                vc.insert(0, by)

        gs = [
            q.groupby(vc).agg(pl.count()).rename({f'PG{i}_p': 'PG', f'PG{i}_f': 'flag', 'count': f'PG{i}'})
            for i, vc in enumerate(vc_subset, 1)
        ]
        return reduce(
            lambda g1, g2: g1.join(g2, on=[by, 'PG', 'flag'] if by else ['PG', 'flag'], how='outer'), gs
        ).with_columns(
            [
                pl.col('PG').cast(pl.Categorical).cat.set_ordering('lexical'),
                pl.col('^PG\d$').fill_null(strategy='zero'),
            ]
        )

    def load_excel(self, fn='Price_is_Right_Frequency.xlsx'):
        self.excel_fp = self.load_func(fn)

    def save_excel(self, fn='Price_is_Right_Frequency.xlsx'):
        self.save_func(self.excel_fp, fn)
        self._reset_excel()

    def _reset_excel(self):
        self.excel_fp.seek(0)

    def load_pickle(self, fn='df_dict.pickle'):
        with self.load_func(fn) as f:
            self._df_dict = pickle.load(f)

    def update(self, prodNumber, pgps, append, airdate, intended_date, notes):
        if not self.excel_fp:
            self.load_excel()

        f_book = openpyxl.load_workbook(self.excel_fp)

        isPrimetime = prodNumber.endswith('SP')

        f_sheet = f_book['Calendar']
        EXCEL_COLORS = [f_sheet[f'N{i}'].fill for i in range(2, 10)]
        EMPTY_FILL = f_sheet['A1'].fill
        if isPrimetime:
            f_sheet = f_book['Primetime']

        if not append:
            row_idx, _, retro_ts = (
                self._df_dict['primetime' if isPrimetime else 'daytime']
                .select(pl.col('^(PROD|INT. DATE)$'))
                .with_row_count()
                .row(by_predicate=pl.col('PROD') == prodNumber)
            )

            idx = 2 + row_idx
            if not isPrimetime:
                idx += self._df_dict['unaired'].filter(pl.col('INT. DATE') < retro_ts).height
        else:
            if isPrimetime:
                idx = 2 + self._df_dict['primetime'].height
            else:
                sorted_index = [
                    SORT_PROD(i) for i in self._df_dict['daytime'].to_series().to_list() if i not in UNAIRED_DUPLICATES
                ]
                idx = (
                    2
                    + len(UNAIRED_DUPLICATES)
                    + self._df_dict['unaired'].height
                    + bisect.bisect(sorted_index, prodNumber[4] + prodNumber[:4], lo=len(sorted_index) - 250)
                )

            f_sheet.insert_rows(idx)

            for c in 'ABCDEFGHIJKL':
                if isPrimetime and c == 'K':
                    continue
                # going into private attributes is not so kosher, but it works! and the library should expose this publicly!
                f_sheet[f'{c}{idx}']._style = copy(f_sheet[f'{c}{idx-1}']._style)
                f_sheet[f'{c}{idx}'].fill = copy(EMPTY_FILL)
                if f_sheet[f'{c}{idx}'].font.color:
                    f_sheet[f'{c}{idx}'].font.color.rgb = 'FF000000'

            f_sheet[f'A{idx}'] = prodNumber
            if not isPrimetime:
                f_sheet[f'B{idx}'] = CURRENT_SEASON
                # f_sheet[f'C{idx}'] = 1 + f_sheet[f'C{idx-1}'].value
            else:
                f_sheet[f'D{idx}'] = 'TPIR@N – '
            for c in 'BC' if isPrimetime else 'DE':
                f_sheet[f'{c}{idx}'] = airdate

        if pgps:
            for c, pgp in zip('EFGHIJ' if isPrimetime else 'GHIJKL', pgps):
                f_sheet[f'{c}{idx}'] = str(pgp.pg) if append else pgp.pg_str
                if not append or pgp.flag:
                    if pgp.flag == 2 ** _pf_bit('?'):
                        f_sheet[f'{c}{idx}'] = '*' + f_sheet[f'{c}{idx}'].value + '*'
                    else:
                        f_sheet[f'{c}{idx}'].fill = copy(EXCEL_COLORS[int(np.log2(pgp.flag))] if pgp.flag else EMPTY_FILL)
        else:
            assert not append
            if airdate:
                c = 'B' if isPrimetime else 'D'
                f_sheet[f'{c}{idx}'] = airdate
            if intended_date:
                c = 'C' if isPrimetime else 'E'
                f_sheet[f'{c}{idx}'] = intended_date
            # allow for unsetting notes if empty string
            if notes is not None:
                c = 'D' if isPrimetime else 'F'
                f_sheet[f'{c}{idx}'] = notes

        self.excel_fp = io.BytesIO()
        f_book.save(self.excel_fp)
        self._reset_excel()

    @staticmethod
    def _parse_sheet(workbook: pd.ExcelFile, sheetName, usecols):
        return workbook.parse(sheet_name=sheetName, index_col=0, usecols=usecols, converters=conv_dict, skipfooter=1)

    @staticmethod
    def _fill_in_flags(dff, fs, col, col_adjust, ec):
        # this method has a lot of (odd-looking) microadjustments since it's run concurrently 3 times / a big factor in loading time
        # d = dff.to_dict('records')
        for i, ep in enumerate(fs.iter_rows(min_row=2, max_row=len(dff.index) + 1, min_col=col, max_col=col + col_adjust)):
            for s, cell in enumerate(ep, -(col_adjust + 1)):
                # try:
                # 	if dff.iloc[i][f'PG{s}'].flag & 2**_pf_bit('?'):	# game uncertainty needs no further checking
                # 		continue
                # except AttributeError:	# nan - no playing here
                # 	continue

                if cell.fill.start_color.rgb in ec:  # and not dff.iloc[i][f'PG{s}'].flag:
                    if not (dff.iat[i, s].flag & 2 ** _pf_bit('?')):
                        dff.iat[i, s].flag |= 2 ** (ec[cell.fill.start_color.rgb])
                # elif color.rgb not in ('FFFFFFFF', '00000000', 'FF7F7F7F'):
                # print(cell)
                # print(color)

                if cell.font.color and cell.font.color.rgb != 'FF000000':
                    try:
                        dff.iat[i, s].flag |= 2 ** (ec[cell.font.color.rgb])
                    except KeyError:
                        dff.iat[i, s].flag |= 2 ** _pf_bit('MDG')

    def _reset_caches(self):
        self.endpoint_sub.cache_clear(self)
        self.lineup_query.cache_clear(self)
        self.concurrence_query.cache_clear(self)
        self.slot_table.cache_clear(self)

    def initialize(self):
        self._df_dict.clear()
        self._reset_caches()
        # print(datetime.datetime.now())
        # openpyxl.load_workbook(self.excel_fp, read_only=True, data_only=True, keep_links=False)
        with pd.ExcelFile(self.excel_fp, engine='openpyxl') as xlsx, ThreadPoolExecutor() as executor:
            # dtype={'S':pd.CategoricalDtype(categories=range(1,CURRENT_SEASON+1), ordered=True)

            df_args = [(xlsx, 'Calendar', 'A:B,D:L'), (xlsx, 'Primetime', 'A:J'), (xlsx, 'Syndication', 'A:E')]
            futures = [executor.submit(ConflictSheet._parse_sheet, *dfa) for dfa in df_args]
            df_daytime, df_primetime, df_syndicated = [f.result() for f in futures]
            # print(datetime.datetime.now())
            df_primetime.dropna(subset=['AIRDATE'], inplace=True)

            # read background colors and adjust
            f_book = xlsx.book
            f_sheet = f_book['Calendar']
            EXCEL_COLORS = {f_sheet[f'N{i}'].fill.start_color.rgb: i - 2 for i in range(2, 10)}

            # read notes
            self.notes = '\n'.join('-' + f_sheet[f'AA{i}'].value for i in range(2, 9))
            # print(datetime.datetime.now())
            df_args = [
                (df_daytime, f_sheet, 7, 5, EXCEL_COLORS),
                (df_primetime, f_book['Primetime'], 5, 5, EXCEL_COLORS),
                (df_syndicated, f_book['Syndication'], 3, 2, EXCEL_COLORS),
            ]
            futures = [executor.submit(ConflictSheet._fill_in_flags, *dfa) for dfa in df_args]
            [f.result() for f in futures]
            # print(datetime.datetime.now())

        self._reset_excel()

        # split into unaired, copy
        df_unaired = df_daytime[df_daytime.apply(lambda r: pd.isnull(r['AIRDATE']), axis=1)].copy(deep=True)
        df_daytime.drop(index=df_unaired.index, inplace=True)

        # raw info
        df_daytime_info = (
            df_daytime.loc[:, 'PG1':'PG6']
            .applymap(lambda v: v.pg if pd.notna(v) else np.nan)
            .join(
                df_daytime.loc[:, 'PG1':'PG6'].applymap(lambda v: v.flag if pd.notna(v) else 2**15).astype('uint16'),
                rsuffix='_f',
            )
            .astype('category')
        )
        df_primetime_info = (
            df_primetime.loc[:, 'PG1':'PG6']
            .applymap(lambda v: v.pg if pd.notna(v) else np.nan)
            .join(
                df_primetime.loc[:, 'PG1':'PG6'].applymap(lambda v: v.flag if pd.notna(v) else 2**15).astype('uint16'),
                rsuffix='_f',
            )
            .astype('category')
        )
        df_unaired_info = (
            df_unaired.loc[:, 'PG1':'PG6']
            .applymap(lambda v: v.pg if pd.notna(v) else np.nan)
            .join(
                df_unaired.loc[:, 'PG1':'PG6'].applymap(lambda v: v.flag if pd.notna(v) else 2**15).astype('uint16'),
                rsuffix='_f',
            )
            .astype('category')
        )
        df_syndicated_info = (
            df_syndicated.loc[:, 'PG1':'PG3']
            .applymap(lambda v: v.pg if pd.notna(v) else np.nan)
            .join(
                df_syndicated.loc[:, 'PG1':'PG3'].applymap(lambda v: v.flag if pd.notna(v) else 2**15).astype('uint16'),
                rsuffix='_f',
            )
            .astype('category')
        )

        # now that all dfs have been built, can replace PGPlaying object with static category strings
        for i in range(1, 7):
            df_daytime[f'PG{i}'] = df_daytime[f'PG{i}'].apply(lambda v: str(v) if pd.notna(v) else np.nan).astype('category')
            df_primetime[f'PG{i}'] = (
                df_primetime[f'PG{i}'].apply(lambda v: str(v) if pd.notna(v) else np.nan).astype('category')
            )
            df_unaired[f'PG{i}'] = df_unaired[f'PG{i}'].apply(lambda v: str(v) if pd.notna(v) else np.nan).astype('category')
            if i <= 3:
                df_syndicated[f'PG{i}'] = (
                    df_syndicated[f'PG{i}'].apply(lambda v: str(v) if pd.notna(v) else np.nan).astype('category')
                )
        df_daytime['NOTES'] = df_daytime['NOTES'].astype('category')
        df_primetime['SPECIAL'] = df_primetime['SPECIAL'].astype('category')

        # build public dict: polars

        for time in ('daytime', 'primetime', 'syndicated', 'unaired'):
            df = eval(f'df_{time}')

            cols = [
                pl.col('PROD'),
                pl.col('^.*DATE$').cast(pl.Date),
                pl.arange(1, len(df) + 1).cast(pl.UInt16).alias('PG_n'),
                pl.col('^PG\d$').cast(pl.Categorical),
            ]

            if time != 'primetime':
                cols.insert(1, pl.col('S').cast(pl.UInt8))
            if time != 'syndicated':
                cols.insert(len(cols) - 2, pl.col('^(NOTES|SPECIAL)$'))

            self._df_dict[time] = (
                pl.from_pandas(df.reset_index())
                .select(cols)
                .hstack(
                    pl.from_pandas(eval(f'df_{time}_info').applymap(lambda v: str(v) if type(v) is PG else v))
                    .select(pl.exclude('S'))
                    .with_columns([pl.col('^PG\d$').cast(pl.Categorical), pl.col('^PG\d_f$').cast(pl.UInt16)])
                    .rename({f'PG{i}': f'PG{i}_p' for i in range(1, 4 if time == 'syndicated' else 7)})
                )
            )

        with io.BytesIO() as dfd_ip:
            pickle.dump(self._df_dict, dfd_ip)
            dfd_ip.seek(0)
            self.save_func(dfd_ip, 'df_dict.pickle')
